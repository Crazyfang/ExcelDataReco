from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors
import logging
import time
import os
from win32com.client import Dispatch
import win32com.client


def generate_logging():
    """
    return:
        The logger output the log message

    """
    # First, generate a logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Second, generate a log handler to write the log file
    rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
    log_path = os.getcwd() + '/Logs/'
    if os.path.exists(log_path):
        pass
    else:
        os.mkdir(log_path)

    log_name = log_path + rq + '.log'
    logfile = log_name
    fh = logging.FileHandler(logfile, mode='w')
    fh.setLevel(logging.DEBUG)

    # Third, define the output format of handler
    formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
    fh.setFormatter(formatter)

    # Four, add the handler into logger
    logger.addHandler(fh)

    return logger


class ProcessData:
    """
    Process excel's data and write the data of the red font to the new sheet
    :param
    file_path: the file's path, can be the relative path or absolute path

    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.max_rows = 1
        self.max_columns = 1
        self.logger = generate_logging()
        self.wb = load_workbook(self.file_path, keep_vba=True, data_only=True)
        self.sheet_names = self.wb.sheetnames
        self.need_shift_data = []
        self.red_place = []
        self.limit = None
        self.range = None

    def read_variable(self, sheet_name=None):
        """
        Read the limit and range value from sheet.If the sheet_name param is none, the function will use the default
        sequence number
        :return
        True/False-The sign of the limit and range value get successfully or not.
        """
        try:
            if sheet_name is None:
                self.limit = self.wb[self.sheet_names[2]].cell(row=2, column=3).value
                self.range = self.wb[self.sheet_names[2]].cell(row=2, column=4).value
            else:
                self.limit = self.wb[sheet_name].cell(row=2, column=3).value
                self.range = self.wb[sheet_name].cell(row=2, column=4).value

            if self.limit is not None and self.range is not None:
                self.logger.info("读取增减额和增减幅成功，增减额:{0}, 增减幅:{1}".format(self.limit, self.range))
                return True
            else:
                self.logger.error("读取增减额和增减幅失败, 失败原因:当前单元格未包含数值, 增减额:{0}, 增减幅:{1}".format(self.limit, self.range))
                return False
        except ValueError as reason:
            self.logger.error(str(reason))
            return False

    def read_month_first_rmb(self):
        """
        月一批人命币Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[3]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[3]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    if row_index != 809:
                        if self.compare_value(list_item[7], self.limit):
                            try:
                                float(list_item[8])
                                list_item[8] = round(list_item[8])
                            except ValueError:
                                pass
                            self.red_place.append([self.get_row_number, 8])
                            if self.compare_value(list_item[8], self.range):
                                self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
                        elif self.compare_value(list_item[8], self.range):
                            try:
                                float(list_item[8])
                                list_item[8] = round(list_item[8])
                            except ValueError:
                                pass
                            self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
                    else:
                        if self.compare_value(list_item[8], self.limit):
                            list_item[8] = round(list_item[8])
                            self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
            self.logger.info('月一批人命币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_month_first_foreign(self):
        """
        月一批外币Sheet表
        :return:
        """
        item_all = []
        item_row = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[4]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[4]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    if self.compare_value(list_item[7], self.limit):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('月一批外币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_month_second_rmb(self):
        """
        月二批人民币Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[5]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[5]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    if self.compare_value(list_item[7], self.limit):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('月二批人民币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_month_second_foreign(self):
        """
        月二批外币Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[6]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[6]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    if self.compare_value(list_item[7], self.limit):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        try:
                            float(list_item[8])
                            list_item[8] = round(list_item[8])
                        except ValueError:
                            pass
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)

            self.logger.info('月二批外币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_3410(self):
        """
        3410 Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[7]].max_row + 1):
                for column in range(1, 15):
                    item_row.append(self.wb[self.sheet_names[7]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                        float(list_item[12])
                        list_item[12] = round(list_item[12])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])

                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[11], self.limit):
                        self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[12], self.range):
                        self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
            self.logger.info('3410表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_3414(self):
        """
        3414 Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[8]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[8]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('3414表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_season_first_rmb(self):
        """
        季一批人民币
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[9]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[9]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('季一批人民币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_season_first_foreign(self):
        """
        季一批外币Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[10]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[10]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('季一批外币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_3301(self):
        """
        3301 Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[11]].max_row + 1):
                for column in range(1, 22):
                    item_row.append(self.wb[self.sheet_names[11]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                        float(list_item[12])
                        list_item[12] = round(list_item[12])
                        float(list_item[16])
                        list_item[16] = round(list_item[16])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        if self.compare_value(list_item[15], self.limit):
                            self.red_place.append([self.get_row_number, 16])
                        if self.compare_value(list_item[16], self.range):
                            self.red_place.append([self.get_row_number, 17])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        if self.compare_value(list_item[15], self.limit):
                            self.red_place.append([self.get_row_number, 16])
                        if self.compare_value(list_item[16], self.range):
                            self.red_place.append([self.get_row_number, 17])

                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[11], self.limit):
                        self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        if self.compare_value(list_item[15], self.limit):
                            self.red_place.append([self.get_row_number, 16])
                        if self.compare_value(list_item[16], self.range):
                            self.red_place.append([self.get_row_number, 17])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[12], self.range):
                        self.red_place.append([self.get_row_number, 13])
                        if self.compare_value(list_item[15], self.limit):
                            self.red_place.append([self.get_row_number, 16])
                        if self.compare_value(list_item[16], self.range):
                            self.red_place.append([self.get_row_number, 17])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[15], self.limit):
                        self.red_place.append([self.get_row_number, 16])
                        if self.compare_value(list_item[16], self.range):
                            self.red_place.append([self.get_row_number, 17])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[16], self.limit):
                        self.red_place.append([self.get_row_number, 17])
                        self.need_shift_data.append(list_item)
            self.logger.info('3414表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_3304(self):
        """
        3304 Sheet
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[12]].max_row + 1):
                for column in range(1, 10):
                    item_row.append(self.wb[self.sheet_names[12]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        self.need_shift_data.append(list_item)
            self.logger.info('3304表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def read_season_sum(self):
        """
        季报含发生额
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb[self.sheet_names[10]].max_row + 1):
                for column in range(1, 15):
                    item_row.append(self.wb[self.sheet_names[10]].cell(row=row, column=column).value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    try:
                        float(list_item[8])
                        list_item[8] = round(list_item[8])
                        float(list_item[12])
                        list_item[12] = round(list_item[12])
                    except ValueError:
                        pass
                    if self.compare_value(list_item[7], self.limit):
                        self.red_place.append([self.get_row_number, 8])
                        if self.compare_value(list_item[8], self.range):
                            self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[8], self.range):
                        self.red_place.append([self.get_row_number, 9])
                        if self.compare_value(list_item[11], self.limit):
                            self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])

                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[11], self.limit):
                        self.red_place.append([self.get_row_number, 12])
                        if self.compare_value(list_item[12], self.range):
                            self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
                    elif self.compare_value(list_item[12], self.range):
                        self.red_place.append([self.get_row_number, 13])
                        self.need_shift_data.append(list_item)
            self.logger.info('季报含发生额表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def main_func(self):
        self.read_month_first_rmb()
        self.read_month_first_foreign()
        self.read_month_second_rmb()
        self.read_month_second_foreign()
        self.read_3410()
        self.read_3414()
        self.read_season_first_rmb()
        self.read_season_first_foreign()
        self.read_3301()
        self.read_3304()
        self.read_season_sum()

    def write_data_to_sheet(self, sheet_name=None):
        if sheet_name is None:
            sheet_name = '汇总'

        ws = self.wb.create_sheet()
        ws.title = sheet_name

        # 写入数据
        for row in range(1, len(self.need_shift_data) + 1):
            for column in range(1, len(self.need_shift_data[row - 1]) + 1):
                ws.cell(row=row, column=column).value = self.need_shift_data[row - 1][column - 1]

        # 改变指定单元格文字
        font = Font(color=colors.RED)
        for item in self.red_place:
            if item:
                ws.cell(row=item[0], column=item[1]).font = font

        self.wb.save(filename=self.file_path)

    @property
    def get_row_number(self):
        return len(self.need_shift_data) + 1

    def test_row(self):
        print(self.wb[self.sheet_names[2]]['E'])

    def read_font_color(self):
        sheet_names = self.wb.sheetnames
        for sheet_name in sheet_names:
            print(sheet_name)
            if sheet_name not in ['说明', '数据库', '增减设定']:
                self.max_rows = self.wb[sheet_name].max_row
                self.max_columns = self.wb[sheet_name].max_column
                self.read_red_font_data_from_sheet(sheet_name)

    def read_red_font_data_from_sheet(self, sheet_name):
        # 获取表头条目
        head_data = []
        head_item_count = 0
        for column in range(1, self.max_columns + 1):
            if column == 1:
                if self.wb[sheet_name].cell(row=1, column=column).value is None:
                    head_data.append(' ')
                else:
                    head_data.append(self.wb[sheet_name].cell(row=1, column=column).value)
            else:
                if self.wb[sheet_name].cell(row=1, column=column).value is None:
                    break
                else:
                    head_data.append(self.wb[sheet_name].cell(row=1, column=column).value)
            head_item_count += 1
        self.need_shift_data.append(head_data)
        self.red_place.append([])

        # 处理表中红字数据
        sheet_data = []
        red = []
        for row in range(2, self.max_rows + 1):
            do_not_sign = False
            sheet_data = []
            red = []
            for column in range(1, head_item_count + 1):
                if self.wb[sheet_name].cell(row=row, column=column).value is None:
                    sheet_data.append(' ')
                else:
                    sheet_data.append(self.wb[sheet_name].cell(row=row, column=column).value)

                if self.wb[sheet_name].cell(row=row, column=column).font.color is not None:
                    if self.wb[sheet_name].cell(row=row, column=column).font.color.rgb == 'FFFF0000':
                        red.append(column)
                        if column != head_item_count:
                            for column_second in range(column + 1, head_item_count + 1):
                                if self.wb[sheet_name].cell(row=row, column=column_second).font.color is not None:
                                    if self.wb[sheet_name].cell(row=row,
                                                                column=column_second).font.color.rgb == 'FFFF0000':
                                        do_not_sign = True
                                    else:
                                        pass
                                else:
                                    pass

                            if not do_not_sign:
                                self.need_shift_data.append(sheet_data)
                                sheet_data = []
                                self.red_place.append(red)
                                red = []
                        else:
                            self.need_shift_data.append(sheet_data)
                            sheet_data = []
                            self.red_place.append(red)
                            red = []

    @staticmethod
    def compare_value(str_value, int_value):
        if str_value is None:
            return False
        else:
            try:
                str_value = int(str_value)
                int_value = int(int_value)
                if str_value >= int_value or str_value <= -int_value:
                    return True
                else:
                    return False
            except ValueError:
                return False

    @property
    def print_data(self):
        return self.need_shift_data


class TestClass:
    def __init__(self, file_path):
        self.file_path = file_path
        self.excel = win32com.client.Dispatch('Excel.Application')
        self.excel.Visible = False
        self.wb = self.excel.Workbooks.Open(self.file_path)
        self.need_shift_data = []
        self.red_place = []
        self.limit = None
        self.range = None
        self.logger = generate_logging()
        self.ws = None

    def read_variable(self, sheet_name=None):
        """
        Read the limit and range value from sheet.If the sheet_name param is none, the function will use the default
        sequence number
        :return
        True/False-The sign of the limit and range value get successfully or not.
        """
        try:
            if sheet_name is None:
                self.limit = self.wb.Worksheets['增减设定'].Cells(2, 3).Value
                self.range = self.wb.Worksheets['增减设定'].Cells(2, 4).Value
            else:
                self.limit = self.wb.Worksheets[sheet_name].Cells(2, 3).Value
                self.range = self.wb.Worksheets[sheet_name].Cells(2, 4).Value

            if self.limit is not None and self.range is not None:
                self.logger.info("读取增减额和增减幅成功，增减额:{0}, 增减幅:{1}".format(self.limit, self.range))
                return True
            else:
                self.logger.error("读取增减额和增减幅失败, 失败原因:当前单元格未包含数值, 增减额:{0}, 增减幅:{1}".format(self.limit, self.range))
                return False
        except ValueError as reason:
            self.logger.error(str(reason))
            return False

    def read_month_first_rmb(self):
        """
        月一批人命币Sheet表
        :return:
        """
        item_row = []
        item_all = []

        try:
            # 先从Sheet表中读取数据至item_all数组
            for row in range(1, self.wb.Worksheets['月一批人民币'].usedrange.rows.count + 1):
                for column in range(1, 10):
                    item_row.append(self.wb.Worksheets['月一批人民币'].Cells(row, column).Value)
                item_all.append(item_row)
                item_row = []

            # 对item_all数组进行条件格式的验证
            for row_index, list_item in enumerate(item_all):
                if row_index == 0:
                    self.red_place.append([])
                    self.need_shift_data.append(list_item)
                else:
                    if row_index != 809:
                        if self.compare_value(list_item[7], self.limit):
                            try:
                                float(list_item[8])
                                list_item[8] = round(list_item[8])
                            except ValueError:
                                pass
                            self.red_place.append([self.get_row_number, 8])
                            if self.compare_value(list_item[8], self.range):
                                self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
                        elif self.compare_value(list_item[8], self.range):
                            try:
                                float(list_item[8])
                                list_item[8] = round(list_item[8])
                            except ValueError:
                                pass
                            self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
                    else:
                        if self.compare_value(list_item[8], self.limit):
                            list_item[8] = round(list_item[8])
                            self.red_place.append([self.get_row_number, 9])
                            self.need_shift_data.append(list_item)
            self.logger.info('月一批人命币表读取成功!')
        except Exception as reason:
            self.logger.error(str(reason))

    def write_data_to_sheet(self, sheet_name=None):
        if sheet_name is None:
            sheet_name = '汇总'

        try:
            self.ws = self.wb.Worksheets[sheet_name]
            self.ws.Rows('1:2000').Delete()
        except Exception as reason:
            self.logger.error(str(reason))
            self.wb.Worksheets.Add().Name = sheet_name
            self.ws = self.wb.Worksheets(sheet_name)

        # 写入数据
        for row in range(1, len(self.need_shift_data) + 1):
            for column in range(1, len(self.need_shift_data[row - 1]) + 1):
                self.ws.Cells(row, column).Value = self.need_shift_data[row - 1][column - 1]

        # 改变指定单元格文字
        for item in self.red_place:
            if item:
                self.ws.Cells(item[0], item[1]).Font.Color = -16776961

        self.wb.Save()
        self.wb.Close()

    def test(self):
        self.wb.Worksheets.Add().Name = '汇总'
        ws = self.wb.Worksheets('汇总')
        ws.Cells(1, 1).Font.Color = -16776961
        ws.Cells(1, 1).Value = '测试'
        self.wb.Save()
        self.wb.Close()

    def main_func(self):
        self.read_variable()
        self.read_month_first_rmb()
        self.write_data_to_sheet()

    @staticmethod
    def compare_value(str_value, int_value):
        if str_value is None:
            return False
        else:
            try:
                str_value = int(str_value)
                int_value = int(int_value)
                if str_value >= int_value or str_value <= -int_value:
                    return True
                else:
                    return False
            except ValueError:
                return False

    @property
    def get_row_number(self):
        return len(self.need_shift_data) + 1


if __name__ == '__main__':
    pd = TestClass('C:\\Users\\Administrator\\Documents\\GitHub\\ExcelDataReco\\（镇海）金融统计报表管理系统（农合用）20200413季报.xlsm')
    pd.main_func()
    # pd = ProcessData('./（镇海）金融统计报表管理系统（农合用）20200413季报.xlsm')
    # pd.test_row()
    # pd.read_variable()
    # pd.main_func()
    # print(len(pd.print_data))
    # print(pd.print_data)
    # pd.write_data_to_sheet()
    # pd.read_font_color()
    # print(pd.print_data)
    # print(pd.red_place)
